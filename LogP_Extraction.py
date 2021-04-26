import csv
from openpyxl import load_workbook
from openpyxl import Workbook

# open LogP File
LogP_File = open("LogP_data.csv")
LogP_Reader = csv.reader(LogP_File)
LogP_Data = list(LogP_Reader)

# Open Compound_Workbook File
Compound_Workbook = load_workbook('target_compounds.xlsx', read_only=True)
Compound_Worksheet = Compound_Workbook["LSER Compounds"]
CAS_Dict = {}

for row in Compound_Worksheet:
    CAS_Line = row[1].value
    CAS_List = CAS_Line.split(",")
    for CAS_Number in CAS_List:
        CAS_Number = CAS_Number.strip()
        if CAS_Number in CAS_Dict:
            print(CAS_Number)
        else:
            CAS_Dict[CAS_Number] = CAS_Number

LogP_Final_wb = Workbook(write_only=True)
LogP_Final_ws = LogP_Final_wb.create_sheet()

LogP_Final_ws.append(["CAS", "Kow", "Kow Data Type", "Kow Reference", "Kowwin", "Error"])
for row in LogP_Data:
    CAS_Line = row[1]
    P_column = row[15]
    Q_column = row[16]
    R_column = row[17]
    T_column = row[19]
    U_column = row[20]
    CAS_List = CAS_Line.split(",")
    for CAS_Number in CAS_List:
        CAS_Number = CAS_Number.strip()
        if CAS_Number in CAS_Dict:
            LogP_Final_ws.append([str(CAS_Number), P_column, Q_column, R_column, T_column, U_column])
    
LogP_Final_wb.save('Final_LogP_Data.xlsx')