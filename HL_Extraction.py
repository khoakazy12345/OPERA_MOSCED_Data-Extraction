import csv
from openpyxl import load_workbook
from openpyxl import Workbook

# open HL File
HL_File = open("HL_data.csv")
HL_Reader = csv.reader(HL_File)
HL_Data = list(HL_Reader)

# Open Compound_Workbook File
Compound_Workbook = load_workbook('target_compounds.xlsx', read_only=True)
Compound_Worksheet = Compound_Workbook["LSER Compounds"]
CAS_Dict = {}

for row in Compound_Worksheet:
    CAS_Line = row[1].value
    CAS_List = CAS_Line.split(",")
    for CAS_Number in CAS_List:
        CAS_Number = CAS_Number.strip()
        if CAS_Number in CAS_Dict:
            print(CAS_Number)
        else:
            CAS_Dict[CAS_Number] = CAS_Number

HL_Final_wb = Workbook(write_only=True)
HL_Final_ws = HL_Final_wb.create_sheet()

HL_Final_ws.append(["CAS", "HL", "HL Temperature", "HL Data Type", "HL Reference", "BondEst", "GroupEst"])
for row in HL_Data:
    CAS_Line = row[1]
    P_column = row[15]
    Q_column = row[16]
    R_column = row[17]
    S_column = row[18]
    T_column = row[19]
    U_column = row[20]
    CAS_List = CAS_Line.split(",")
    for CAS_Number in CAS_List:
        CAS_Number = CAS_Number.strip()
        if CAS_Number in CAS_Dict:
            HL_Final_ws.append([str(CAS_Number), P_column, Q_column, R_column, S_column, T_column, U_column])
    
HL_Final_wb.save('Final_HL_Data.xlsx')