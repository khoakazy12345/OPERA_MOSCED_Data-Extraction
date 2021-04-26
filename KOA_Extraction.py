import csv
from openpyxl import load_workbook
from openpyxl import Workbook

# open KOA File
KOA_File = open("KOA_data.csv")
KOA_Reader = csv.reader(KOA_File)
KOA_Data = list(KOA_Reader)

# Open Compound_Workbook File
Compound_Workbook = load_workbook('target_compounds.xlsx', read_only=True)
Compound_Worksheet = Compound_Workbook["LSER Compounds"]
CAS_Dict = {}

for row in Compound_Worksheet:
    CAS_Line = row[1].value
    CAS_List = CAS_Line.split(",")
    for CAS_Number in CAS_List:
        CAS_Number = CAS_Number.strip()
        if CAS_Number in CAS_Dict:
            print(CAS_Number)
        else:
            CAS_Dict[CAS_Number] = CAS_Number

KOA_Final_wb = Workbook(write_only=True)
KOA_Final_ws = KOA_Final_wb.create_sheet()

KOA_Final_ws.append(["CAS", "KOA", "KOA Temperature", "KOA Data Type", "KOA Reference", "LogKOA", "EstLogKOA", "ErrorLogKOA","KOAest"])
for row in KOA_Data:
    CAS_Line = row[1]
    P_column = row[15]
    Q_column = row[16]
    R_column = row[17]
    S_column = row[18]
    T_column = row[19]
    U_column = row[20]
    V_column = row[21]
    W_column = row[22]
    CAS_List = CAS_Line.split(",")
    for CAS_Number in CAS_List:
        CAS_Number = CAS_Number.strip()
        if CAS_Number in CAS_Dict:
            KOA_Final_ws.append([str(CAS_Number), P_column, Q_column, R_column, S_column, T_column, U_column, V_column, W_column])
    
KOA_Final_wb.save('Final_KOA_Data.xlsx')